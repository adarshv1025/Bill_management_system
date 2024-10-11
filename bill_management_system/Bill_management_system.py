from tkinter import *
import random
from PIL import Image,ImageTk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook #used to create specific programming excel file
import pathlib
import os

root=Tk() #class is already present we r creating object of this cls
root.title("Bill management system") #for title
root.geometry('1466x700') #for geometry
a=Label(root,text="Bill management system",font=('times new roman',30,'bold'),bg='black',width='60',fg='gold',bd=12,relief=RIDGE).grid(row=0)#labeling
#fg=forecolour for text
#bd=border and relief for styling the border


def ex():
    root.destroy()

file=pathlib.Path('bill.xlsx')#identify the file we r giving this
if file.exists():
    pass
else:
    file=Workbook()#to create on top level container
    sheet=file.active  #selects the first sheet
    sheet['A1']="Bill NUMBER"
    sheet['B1']="DOSA"
    sheet['C1']="COFFEE"
    sheet['D1']="TEA"
    sheet['E1']="BIRYANI"
    sheet['F1']="CHAPATI"
    sheet['G1']="Puri"
    sheet['H1']="TOTAL PRICE"
    

    file.save('bill.xlsx')

def save():
    if o==35:
        messagebox.showerror("ERROR","Enter first")
    else:
        file=openpyxl.load_workbook('bill.xlsx')#used to read or write in the excel file  #workbook Workbook is the top-level container for all document information
        sheet=file.active     #to load_workbook used to access that file
        sheet.cell(column=1,row=sheet.max_row+1,value=a)
        sheet.cell(column=2,row=sheet.max_row,value=k1)
        sheet.cell(column=3,row=sheet.max_row,value=k2)
        sheet.cell(column=4,row=sheet.max_row,value=k3)
        sheet.cell(column=5,row=sheet.max_row,value=k4)
        sheet.cell(column=6,row=sheet.max_row,value=k5)
        sheet.cell(column=7,row=sheet.max_row,value=k6)
        sheet.cell(column=8,row=sheet.max_row,value=o)
        file.save('bill.xlsx')
        messagebox.showinfo("info","BILL SAVED SUCCESSFULLY")


    


#import qrcode

#img = qrcode.make("THANK YOU")
#img.save('QRCODE.jpg')



def generate_bill():


    def billexit():
        BILL.destroy()

    if o==35:
        messagebox.showerror('ERROR',"GENERATE TOTAL")
    elif a==None:
        messagebox.showerror('ERROR',"GENERATE BILL NUMBER")
    else:
        BILL=Tk()
        BILL.title("BILL")
        BILL.geometry('500x650')
        x=Label(BILL,text="BILL",font=('times new roman',18,'bold'),bg='black',width='34',fg='gold',bd=12,relief=RIDGE).grid(row=0)


        T = Text(BILL, height = 50, width = 62)
        T.grid(row=1,column=0)
        T.insert(END,'************************WELCOME CUSTOMER**********************\n')
        T.insert(END,f'\nBILL NUMBER: {a}\n')
        T.insert(END,f'\n==============================================================')
        T.insert(END,f'\nPRODUCT\t\t\tQUANTITY\t\t\tPRICE')
        T.insert(END,f'\n==============================================================')

        if a1!=0:
            T.insert(END,f'\nDOSA\t\t\t{a1}\t\t\t{k1}.Rs')
        if a2!=0:
            T.insert(END,f'\nCOFFEE\t\t\t{a2}\t\t\t{k2}.Rs')
        if a3!=0:
            T.insert(END,f'\nTEA\t\t\t{a3}\t\t\t{k3}.Rs')
        if a4!=0:
            T.insert(END,f'\nBIRYANI\t\t\t{a4}\t\t\t{k4}.Rs')
        if a5!=0:
            T.insert(END,f'\nCHAPATI\t\t\t{a5}\t\t\t{k5}.Rs')
        if a6!=0:
            T.insert(END,f'\nPURI\t\t\t{a6}\t\t\t{k6}.Rs')

        T.insert(END,f'\n==============================================================')
        T.insert(END,'\nGST TAX\t\t\t\t15.Rs')
        T.insert(END,'\nSEERVICE CHARGE\t\t\t\t20.Rs')
        T.insert(END,f'\n==============================================================')
        T.insert(END,f'\nTOTAL PRICE\t\t\t\t{o}.Rs')
        T.insert(END,f'\n==============================================================')
        T.insert(END,f'\n\t\t\tTHANK YOU')


        Total.destroy()

        
        def add():
            if not os.path.exists('bills'):
                os.mkdir('bills')
            bill_content=T.get(1.0,END)
            f=open(f'bills/ {a}.txt','w')
            f.write(bill_content)
            f.close()
            messagebox.showinfo("info","BILL ADDED SUCCESSFULLY")
            BILL.destroy()

        btn_add=Button(BILL,bd=5,fg="black",bg="pink",font=("times new roman",16,"bold"),width=10,text="ADD BILL",command=add)
        btn_add.place(x=100,y=500)

        btn_billexit=Button(BILL,bd=5,fg="black",bg="pink",font=("times new roman",16,"bold"),width=10,text="EXIT",command=billexit)
        btn_billexit.place(x=300,y=500)

        
        BILL.mainloop()


def search_bill():
    search=Tk()
    search.title("SEARCH BILL")
    search.geometry('500x650')

    def searchexit():
        search.destroy()

    def sea():
        global a
        for i in os.listdir('bills/'): #To get list of all file directries in that folder
            a=entry_search.get()
            b=int(a)
            c=i.split('.')
            d=int(c[0])
            if d==b:#To check the given input is present or not
                f=open(f'bills/{i}','r')#To open it in Read mode
                Z.delete(1.0,END)#Delete Previous Bill
                for data in f:          #Insert Each Line into Text Area
                    Z.insert(END,data)
                f.close()
                break
        else: #If File not Present shows Error
                messagebox.showerror('Error','Invalid Bill Number')


    

    Z = Text(search, height = 50, width = 62)
    Z.place(x=0,y=70)

    


    btn_sea=Button(search,bd=5,fg="black",bg="lightgreen",font=("times new roman",16,"bold"),width=10,text="SEARCH",command=sea)
    btn_sea.place(x=20,y=10)

    
    entry_search=Entry(search,font=("aria",16,'bold'),bd=6,width=20,bg="pink")
    entry_search.place(x=200,y=15)

    btn_exit=Button(search,bd=5,fg="black",bg="yellow",font=("times new roman",16,"bold"),width=10,text="EXIT",command=searchexit)
    btn_exit.place(x=200,y=550)

    search.mainloop()
    


    
def bill():
    global a
    a=random.randint(10000,999999)
    billnum.set(a)


def Reset():  #for reseting all
    entry_Dosa.delete(0,END)
    entry_Coffee.delete(0,END)
    entry_Tea.delete(0,END)
    entry_Biryani.delete(0,END)
    entry_Chapati.delete(0,END)
    entry_Puri.delete(0,END)



def qr_code():
    
    from PIL import Image,ImageTk
    global photo,image,img

    
    import qrcode

    img = qrcode.make(f'YOUR TOTAL IS {o} RS \n THANK YOU VISIT AGAIN \nTHIS IS POEM TO MAKE YOUR DAY AMAZING \n ಗುಲಾಬಿ ಹೂವು ನೀ ಕಂಡಂತೆ\n ಸಾಗರದ ಅಲೆಯು ನಿನಗೆ  ಬಡಿದಂತ\nಆ ಲೋಕ ಸುಂದರ\n ಬಾ ಬಂದು ನೀ ಕುಳಿತುಕೋ ಈ ಹೂವಿನ ತೋಟದಲಿ\n ಬಾ ಬಂದು ಸವಿ ನಿನಗಾಗ್ಲಿಯೇ ಇರುವ ಈ ಆಹಾರ..\nಮಧುರ ಬಾಳು ನಮ್ಮ ಸಂಸ್ಕೃತಿ ದೊಂದಿಗೆ')
    img.save('QRCODE.jpg')

    qr=Toplevel()
    qr.title("QR CODE")
    qr.geometry('500x510')
    x=Label(qr,text="QR CODE",font=('times new roman',18,'bold'),bg='black',width='34',fg='gold',bd=12,relief=RIDGE).grid(row=0)
   
    photo=Image.open("QRCODE.jpg")
    photo=photo.resize((400,400))
    img=ImageTk.PhotoImage(photo)
    label=Label(qr,image=img).place(x=50,y=60)

    qr.mainloop()



def total_price():
    global o,Total
    Total=Tk()
    Total.title("TOTAL PRICE")
    Total.geometry('500x510')
    x=Label(Total,text="TOTAL",font=('times new roman',18,'bold'),bg='black',width='34',fg='gold',bd=12,relief=RIDGE).grid(row=0)


    f4=Frame(Total,bg="black",highlightbackground="white",highlightthickness="1",width="500",height="465",bd=7,relief=RIDGE)
    f4.place(x=0,y=48)


    totalprice=StringVar()

    global a1
    global a2
    global a3
    global a4
    global a5
    global a6

    try : a1=int(Dosa.get())
    except : a1=0 

    try : a2=int(Coffee.get())
    except : a2=0

    try : a3=int(Tea.get())
    except : a3=0

    try : a4=int(Biryani.get())
    except : a4=0

    try : a5=int(Chapati.get())
    except : a5=0

    try : a6=int(Puri.get())
    except : a6=0

    global k1
    global k2
    global k3
    global k4
    global k5
    global k6

    k1=a1*50
    k2=a2*15
    k3=a3*15
    k4=a4*100
    k5=a5*30
    k6=a6*40

    gst=int(15)

    service=int(20)

    u=gst+service+k1+k2+k3+k4+k5+k6
    o=int(u)

    lbl_GST=Label(f4,font=("times new roman",18,'bold'),text='GST TAX',fg="white",bg="black")
    lbl_GST.place(x=30,y=60)

    entry_GST=Entry(Total,font=("aria",16,'bold'),bd=6,width=10,bg="pink")
    entry_GST.place(x=300,y=105)
    entry_GST.insert(0,f'{gst} Rs')

    lbl_service=Label(f4,font=("times new roman",18,'bold'),text='SERVICE CHARGE',fg="white",bg="black")
    lbl_service.place(x=30,y=180)

    entry_service=Entry(Total,font=("aria",16,'bold'),bd=6,width=10,bg="pink")
    entry_service.place(x=300,y=225)
    entry_service.insert(0,f'{service} Rs')

    lbl_Total=Label(f4,font=("times new roman",18,'bold'),text='TOTAL PRICE',fg="white",bg="black")
    lbl_Total.place(x=30,y=300)

    entry_Total=Entry(Total,font=("aria",16,'bold'),bd=6,width=10,bg="pink",textvariable=totalprice)
    entry_Total.place(x=300,y=345)
    entry_Total.insert(0,f'{o} Rs')

    Total.mainloop()
    








    

 
billnum=StringVar()

bill_number=Frame(root,bg='gray20',relief=RIDGE).grid()
billno=Label(bill_number,text='BILL NUMBER  :  ',font=('times new roman',20,'bold'),bg='black',width='90',fg='gold',bd=12,relief=RIDGE).grid(row=1)


billentry=Entry(bill_number,textvariable=billnum,font=('arial',15),width='18')
billentry.place(x=850,y=84)

generatebutton=Button(bill_number,text="Generate",command=bill,font=('arial',10,'bold'))
generatebutton.place(x=1060,y=84)

f=Frame(root,bg="black",highlightbackground="white",highlightthickness="1",width="700",height="500",bd=7,relief=RIDGE)
f.place(x=0,y=127)

Label(f,text=" MENU ",font=("Gabriola",40,"bold"),fg="white",bg="black").place(x=250,y=0)

Label(f,font=("Gabriola",30,'bold'),text="Dosa ------- 50 rs per plate",fg="white",bg="black").place(x=150,y=70)
Label(f,font=("Gabriola",30,'bold'),text="Coffee------ 15 rs per glass",fg="white",bg="black").place(x=150,y=130)
Label(f,font=("Gabriola",30,'bold'),text="Tea --------  15 rs per glass",fg="white",bg="black").place(x=150,y=190)
Label(f,font=("Gabriola",30,'bold'),text="Biryani-----  100 rs per plate",fg="white",bg="black").place(x=150,y=250)
Label(f,font=("Gabriola",30,'bold'),text="Chapati----- 30 rs per plate",fg="white",bg="black").place(x=150,y=310)
Label(f,font=("Gabriola",30,'bold'),text="Puri  -------  40 rs per plate",fg="white",bg="black").place(x=150,y=370)
#Label(f,font=("Gabriola",35,'bold'),text="Idli  - -------  30 rs per plate",fg="white",bg="black").place(x=150,y=430)

f1=Frame(root,bg="black",highlightbackground="white",highlightthickness="1",width="770",height="500",bd=7,relief=RIDGE)
f1.place(x=697,y=127)

Dosa=StringVar()
Coffee=StringVar()
Tea=StringVar()
Biryani=StringVar()
Chapati=StringVar()
Puri=StringVar()

#dosa labeling
lbl_Dosa=Label(f1,font=("times new roman",30,'bold'),text='Dosa',width=12,fg="white",bg="black")
lbl_Dosa.place(x=100,y=20)

#dosa entering
entry_Dosa=Entry(f1,font=("aria",20,'bold'),textvariable=Dosa,bd=6,width=10,bg="gold")
entry_Dosa.place(x=400,y=20)
entry_Dosa.insert(0,0)

#coffee labeling
lbl_Coffee=Label(f1,font=("times new roman",30,'bold'),text='Coffee',width=12,fg="white",bg="black")
lbl_Coffee.place(x=100,y=80)

#coffee entering
entry_Coffee=Entry(f1,font=("aria",20,'bold'),textvariable=Coffee,bd=6,width=10,bg="gold")
entry_Coffee.place(x=400,y=80)
entry_Coffee.insert(0,0)

#tea labeling
lbl_Tea=Label(f1,font=("times new roman",30,'bold'),text='Tea',width=12,fg="white",bg="black")
lbl_Tea.place(x=100,y=140)

#tea entering
entry_Tea=Entry(f1,font=("aria",20,'bold'),textvariable=Tea,bd=6,width=10,bg="gold")
entry_Tea.place(x=400,y=140)
entry_Tea.insert(0,0)


#biryanin labeling
lbl_Biryani=Label(f1,font=("times new roman",30,'bold'),text='Biryani',width=12,fg="white",bg="black")
lbl_Biryani.place(x=100,y=200)


#biryani entering
entry_Biryani=Entry(f1,font=("aria",20,'bold'),textvariable=Biryani,bd=6,width=10,bg="gold")
entry_Biryani.place(x=400,y=200)               
entry_Biryani.insert(0,0)


#chapati labeling
lbl_Chapati=Label(f1,font=("times new roman",30,'bold'),text='Chapati',width=12,fg="white",bg="black")
lbl_Chapati.place(x=100,y=260)

#chapati entering
entry_Chapati=Entry(f1,font=("aria",20,'bold'),textvariable=Chapati,bd=6,width=10,bg="gold")
entry_Chapati.place(x=400,y=260)
entry_Chapati.insert(0,0)

#puri labeling
lbl_Puri=Label(f1,font=("times new roman",30,'bold'),text='Puri',width=12,fg="white",bg="black")
lbl_Puri.place(x=100,y=320)

#puri entering
entry_Puri=Entry(f1,font=("aria",20,'bold'),textvariable=Puri,bd=6,width=10,bg="gold")
entry_Puri.place(x=400,y=320)
entry_Puri.insert(0,0)

#reset
btn_reset=Button(f1,bd=5,fg="black",bg="lightgreen",font=("times new roman",16,"bold"),width=10,text="RESET",command=Reset)
btn_reset.place(x=330,y=400)

f3=Frame(root,bg="black",highlightbackground="white",highlightthickness="1",width="1467",height="80",bd=7,relief=RIDGE)
f3.place(x=0,y=625)

btn_total=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="TOTAL",command=total_price)
btn_total.place(x=10,y=10)

btn_bill=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="GENERATE BILL",command=generate_bill)
btn_bill.place(x=250,y=10)

btn_qrcode=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="QR CODE",command=qr_code)
btn_qrcode.place(x=485,y=10)

btn_save=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="SAVE",command=save)
btn_save.place(x=720,y=10)

btn_search=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="SEARCH",command=search_bill)
btn_search.place(x=955,y=10)

btn_exit=Button(f3,bd=5,fg="black",bg="white",font=("times new roman",16,"bold"),width=17,text="EXIT",command=ex)
btn_exit.place(x=1210,y=10)






root.mainloop()  #to show the windows

